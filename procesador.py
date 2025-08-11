import re
from io import BytesIO
from datetime import datetime, time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# xlrd es opcional, pero da soporte a .xls binario real
try:
    import xlrd
    HAS_XLRD = True
except Exception:
    HAS_XLRD = False

# =================== FERIADOS (si aplica a reglas futuras) ===================
FERIADOS_2025 = [
    "2025-01-01", "2025-04-18", "2025-04-19", "2025-05-01",
    "2025-05-21", "2025-06-29", "2025-07-16", "2025-08-15",
    "2025-09-18", "2025-09-19", "2025-10-12", "2025-10-31",
    "2025-11-01", "2025-12-08", "2025-12-25"
]
FERIADOS_2025 = [datetime.strptime(d, "%Y-%m-%d").date() for d in FERIADOS_2025]

# =================== AUXILIARES GENERALES ===================

def normalizar_fecha(fecha):
    """Convierte 'dd-mm-aaaa' o 'dd/mm/aaaa' o datetime a date."""
    if isinstance(fecha, datetime):
        return fecha.date()
    try:
        s = str(fecha).strip()
        if "-" in s:
            return datetime.strptime(s, "%d-%m-%Y").date()
        if "/" in s:
            return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        return None
    return None

def normalizar_fecha_xls(valor, datemode):
    """
    Convierte un valor de fecha proveniente de hoja .xls (xlrd):
    - numérico Excel -> date
    - 'dd-mm-aaaa' o 'dd/mm/aaaa' -> date
    - datetime -> date
    """
    if isinstance(valor, datetime):
        return valor.date()

    # numérico Excel
    if isinstance(valor, (int, float)) and HAS_XLRD:
        try:
            return xlrd.xldate_as_datetime(valor, datemode).date()
        except Exception:
            pass

    # string
    try:
        s = str(valor).strip()
        if "-" in s:
            return datetime.strptime(s, "%d-%m-%Y").date()
        if "/" in s:
            return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        return None

    return None

def minutos_a_hhmm(minutos):
    return f"{int(minutos) // 60:02}:{int(minutos) % 60:02}"

def convertir_a_hhmm(horas):
    minutos = int(round(float(horas) * 60))
    return minutos_a_hhmm(minutos)

def obtener_horario_turno(turno, dia_semana):
    """
    Extrae horas del string de turno. Ejemplos:
    'Lu-Ju 08:00-17:00; Vi 08:00-16:00' o cualquier texto que contenga hh:mm.
    Para Lu-Ju (0..3) usa el primer par, para Viernes (4) el segundo si existe.
    """
    horas = re.findall(r"\d{1,2}:\d{2}", turno or "")
    if len(horas) < 2:
        return None, None
    # Lu-Ju
    if dia_semana in [0, 1, 2, 3]:
        return horas[0], horas[1]
    # Viernes
    if dia_semana == 4 and len(horas) >= 4:
        return horas[2], horas[3]
    return None, None

def calcular_atraso(entrada, fecha, turno):
    """
    entrada: 'HH:MM:SS' o '-' ; fecha: 'dd-mm-aaaa' o datetime/date
    turno: string con horarios.
    """
    if not entrada or entrada == "-":
        return 0
    f = normalizar_fecha(fecha)
    if not f:
        return 0
    try:
        ent_t = datetime.strptime(str(entrada), "%H:%M:%S").time()
    except Exception:
        return 0
    dia = f.weekday()
    ini_str, _ = obtener_horario_turno(turno, dia)
    if not ini_str:
        return 0
    hora_inicio = datetime.strptime(ini_str, "%H:%M").time()
    if ent_t > hora_inicio:
        return int((datetime.combine(f, ent_t) - datetime.combine(f, hora_inicio)).total_seconds() / 60)
    return 0

def calcular_horas_extras(entrada, salida, fecha, turno, descripcion):
    """
    Devuelve (horas_50, horas_25) como floats según reglas:
    - Antes de la jornada: <07:00 -> 50%, >=07:00 hasta inicio -> 25%
    - Después de la jornada: hasta 21:00 -> 25%, posterior -> 50%
    - Se descartan bloques <= 30 min.
    """
    if not entrada or not salida or entrada == "-" or salida == "-":
        return 0, 0
    if str(descripcion or "").lower() in ("ausente",) or "libre" in str(descripcion or "").lower():
        return 0, 0

    f = normalizar_fecha(fecha)
    if not f:
        return 0, 0
    try:
        e_dt = datetime.strptime(str(entrada), "%H:%M:%S")
        s_dt = datetime.strptime(str(salida), "%H:%M:%S")
    except Exception:
        return 0, 0
    if s_dt < e_dt:
        s_dt += pd.Timedelta(days=1)

    dia = f.weekday()
    ini_str, fin_str = obtener_horario_turno(turno, dia)
    if not ini_str or not fin_str:
        total_min = (s_dt - e_dt).total_seconds() / 60
        return (total_min / 60, 0) if total_min > 30 else (0, 0)

    h_ini = datetime.combine(f, datetime.strptime(ini_str, "%H:%M").time())
    h_fin = datetime.combine(f, datetime.strptime(fin_str, "%H:%M").time())

    min50 = 0
    min25 = 0

    if e_dt < h_ini:
        if e_dt.time() < time(7, 0):
            min50 += (min(s_dt, h_ini) - e_dt).total_seconds() / 60
        else:
            min25 += (min(s_dt, h_ini) - e_dt).total_seconds() / 60

    if s_dt > h_fin:
        if s_dt > datetime.combine(f, time(21, 0)):
            min25 += max(0, (datetime.combine(f, time(21, 0)) - h_fin).total_seconds() / 60)
            min50 += (s_dt - datetime.combine(f, time(21, 0))).total_seconds() / 60
        else:
            min25 += (s_dt - h_fin).total_seconds() / 60

    h50 = round(min50 / 60, 2) if min50 > 30 else 0
    h25 = round(min25 / 60, 2) if min25 > 30 else 0
    return h50, h25

# =================== EXPORTAR RESULTADOS ===================

def armar_resultados(registros, meta):
    """
    registros: lista de dicts con llaves:
      Fecha (str 'dd-mm-aaaa' preferente), Entrada "HH:MM:SS", Salida "HH:MM:SS", Descripcion (str)
    meta: dict con Funcionario, Rut, Organigrama, Turno, Periodo
    """
    detalle = []
    resumen = {}

    for r in registros:
        fecha = r.get("Fecha")
        entrada = r.get("Entrada")
        salida = r.get("Salida")
        descr = r.get("Descripcion", "")

        atraso_min = calcular_atraso(entrada, fecha, meta.get("Turno"))
        h50, h25 = calcular_horas_extras(entrada, salida, fecha, meta.get("Turno"), descr)

        detalle.append([
            meta.get("Funcionario", ""),
            meta.get("Rut", ""),
            meta.get("Organigrama", ""),
            meta.get("Turno", ""),
            meta.get("Periodo", ""),
            fecha, entrada, salida,
            minutos_a_hhmm(atraso_min),
            convertir_a_hhmm(h50),
            convertir_a_hhmm(h25),
            descr
        ])

        f = meta.get("Funcionario", "")
        if f not in resumen:
            resumen[f] = {
                "Rut": meta.get("Rut", ""),
                "Organigrama": meta.get("Organigrama", ""),
                "Turno": meta.get("Turno", ""),
                "Periodo": meta.get("Periodo", ""),
                "Total 50%": 0, "Total 25%": 0, "Total Atraso": 0
            }
        resumen[f]["Total 50%"] += h50
        resumen[f]["Total 25%"] += h25
        resumen[f]["Total Atraso"] += atraso_min

    df_detalle = pd.DataFrame(detalle, columns=[
        "Funcionario", "Rut", "Organigrama", "Turno", "Periodo",
        "Fecha", "Entrada", "Salida", "Atraso (hh:mm)", "50%", "25%", "Descripción"
    ])

    df_resumen = pd.DataFrame([
        [
            f, r["Rut"], r["Organigrama"], r["Turno"], r["Periodo"],
            convertir_a_hhmm(r["Total 50%"]),
            convertir_a_hhmm(r["Total 25%"]),
            minutos_a_hhmm(r["Total Atraso"]),
            convertir_a_hhmm(r["Total 50%"] + r["Total 25%"])
        ]
        for f, r in resumen.items()
    ], columns=["Funcionario", "Rut", "Organigrama", "Turno", "Periodo",
                "Total 50%", "Total 25%", "Total Atraso", "Total Horas"])

    # Exportar + colores
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_detalle.to_excel(writer, sheet_name="Detalle Diario", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
    out.seek(0)

    wb = load_workbook(out)
    ws = wb["Detalle Diario"]
    fill_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Ausente
    fill_amarillo = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Falta entrada/salida o '-'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
        descripcion = str(row[11].value or "")
        entrada = str(row[6].value or "").strip()
        salida  = str(row[7].value or "").strip()
        if "Ausente" in descripcion:
            for c in row:
                c.fill = fill_rojo
        elif "Falta Entrada" in descripcion or "Falta Salida" in descripcion or entrada == "-" or salida == "-":
            for c in row:
                c.fill = fill_amarillo

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# =================== SOPORTE EXCEL REAL (.xlsx y .xls) ===================

def procesar_excel(file_stream: BytesIO) -> BytesIO:
    """
    Intenta procesar como:
      1) .xlsx / .xlsm con openpyxl
      2) .xls binario 97-2003 con xlrd (si disponible)
    Si no detecta layout esperado, levanta excepción para que app.py informe.
    """
    data = file_stream.read()
    # --------- Intento 1: .xlsx (openpyxl) ----------
    try:
        wb = load_workbook(filename=BytesIO(data), data_only=True)
        sh = wb.active

        registros = []
        meta = {"Funcionario": "", "Rut": "", "Organigrama": "", "Turno": "", "Periodo": ""}

        fila = 1
        max_row = sh.max_row
        while fila <= max_row:
            celda = str(sh.cell(row=fila, column=1).value or "").strip().lower()
            if celda.startswith("funcionario"):
                meta["Funcionario"] = str(sh.cell(row=fila, column=2).value or "").strip(": ")
                meta["Rut"] = str(sh.cell(row=fila + 1, column=2).value or "").strip(": ")
                meta["Organigrama"] = str(sh.cell(row=fila + 2, column=2).value or "").strip(": ")
                meta["Turno"] = str(sh.cell(row=fila + 3, column=2).value or "").strip(": ")
                meta["Periodo"] = str(sh.cell(row=fila + 4, column=2).value or "").strip(": ")
                fila += 6
                continue

            if celda == "dia":
                fila += 1
                while fila <= max_row and sh.cell(row=fila, column=1).value:
                    dtext = str(sh.cell(row=fila, column=1).value or "").strip().lower()
                    if dtext == "totales":
                        fila += 1
                        continue
                    if dtext.startswith("funcionario") or dtext == "none":
                        break

                    fecha = sh.cell(row=fila, column=2).value
                    entrada = sh.cell(row=fila, column=3).value
                    salida = sh.cell(row=fila, column=4).value
                    descr = str(sh.cell(row=fila, column=6).value or "").strip()

                    # Normalizar a strings esperadas
                    fecha_str = ""
                    try:
                        f = pd.to_datetime(fecha, dayfirst=True, errors="coerce")
                        if pd.notna(f):
                            fecha_str = f.strftime("%d-%m-%Y")
                    except Exception:
                        pass
                    entrada_str = _to_hms(entrada)
                    salida_str  = _to_hms(salida)

                    registros.append({
                        "Fecha": fecha_str or str(fecha),
                        "Entrada": entrada_str,
                        "Salida": salida_str,
                        "Descripcion": descr
                    })
                    fila += 1
            else:
                fila += 1

        if not registros:
            raise Exception("No se detectó el layout esperado en Excel (xlsx).")
        return armar_resultados(registros, meta)

    except Exception:
        # --------- Intento 2: .xls (xlrd) ----------
        if not HAS_XLRD:
            raise Exception("Intento de leer .xls sin xlrd instalado.")
        try:
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)

            def get(r, c):
                """ acceso 1-based tipo get(fila, columna) """
                try:
                    return sheet.cell_value(r - 1, c - 1)
                except Exception:
                    return None

            registros = []
            meta = {"Funcionario": "", "Rut": "", "Organigrama": "", "Turno": "", "Periodo": ""}

            fila = 1
            max_row = sheet.nrows
            while fila <= max_row:
                celda = str(get(fila, 1)).strip().lower()
                if celda.startswith("funcionario"):
                    meta["Funcionario"] = str(get(fila, 2)).strip(": ")
                    meta["Rut"] = str(get(fila + 1, 2)).strip(": ")
                    meta["Organigrama"] = str(get(fila + 2, 2)).strip(": ")
                    meta["Turno"] = str(get(fila + 3, 2)).strip(": ")
                    meta["Periodo"] = str(get(fila + 4, 2)).strip(": ")
                    fila += 6
                    continue

                if celda == "dia":
                    fila += 1
                    while fila <= max_row and str(get(fila, 1)).strip():
                        dia_text = str(get(fila, 1)).strip().lower()
                        if dia_text == "totales":
                            fila += 1
                            continue
                        if dia_text.startswith("funcionario") or dia_text == "none":
                            break

                        fecha_raw = get(fila, 2)
                        entrada_raw = get(fila, 3)
                        salida_raw  = get(fila, 4)
                        descripcion = str(get(fila, 6) or "").strip()

                        # fecha a dd-mm-aaaa
                        fecha_date = normalizar_fecha_xls(fecha_raw, book.datemode)
                        fecha_str = fecha_date.strftime("%d-%m-%Y") if fecha_date else str(fecha_raw)

                        registros.append({
                            "Fecha": fecha_str,
                            "Entrada": _to_hms(entrada_raw),
                            "Salida": _to_hms(salida_raw),
                            "Descripcion": descripcion
                        })
                        fila += 1
                else:
                    fila += 1

            if not registros:
                raise Exception("No se detectó el layout esperado en Excel (xls).")
            return armar_resultados(registros, meta)

        except Exception as e2:
            raise Exception(f"No se pudo leer Excel real (.xlsx/.xls): {e2}")

# =================== .XLS “HTML” (PLATAFORMA) ===================

def _decode_html_with_fallback(html_bytes):
    """
    Intenta decodificar HTML con varias codificaciones comunes
    (utf-8, latin-1, cp1252...). Si falla, ignora errores.
    """
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252", "iso-8859-1", "windows-1252"]
    for enc in encodings:
        try:
            return html_bytes.decode(enc)
        except Exception:
            continue
    return html_bytes.decode("latin-1", errors="ignore")

def detectar_html_y_procesar(html_bytes):
    """
    Lee las tablas del HTML (archivo con extensión .xls generado por la plataforma),
    intenta mapear columnas y metadatos esperados. Si no puede, devuelve
    un XLSX con las tablas crudas para no cortar el flujo.
    """
    # 1) Decodificar con fallback para soportar acentos/ñ
    html_text = _decode_html_with_fallback(html_bytes)

    # 2) Intentar leer tablas
    try:
        dfs = pd.read_html(html_text, flavor="bs4")  # requiere beautifulsoup4, lxml, html5lib
    except Exception as e:
        raise Exception(f"No se pudieron leer tablas HTML: {e}")

    if not dfs:
        raise Exception("No se encontraron tablas en el archivo HTML.")

    # 3) Extraer metadatos (Funcionario, Rut, Organigrama, Turno, Periodo)
    meta = {"Funcionario": "", "Rut": "", "Organigrama": "", "Turno": "", "Periodo": ""}
    for df in dfs:
        # examina primeras filas buscando pares clave/valor
        to_check = min(len(df), 12)
        for i in range(to_check):
            try:
                fila = [str(x).strip() for x in list(df.iloc[i].values)]
            except Exception:
                continue
            if len(fila) >= 2:
                k = fila[0].lower()
                v = fila[1]
                if "funcionario" in k:
                    meta["Funcionario"] = v
                elif k == "rut" or "rut" in k:
                    meta["Rut"] = v
                elif "organigrama" in k:
                    meta["Organigrama"] = v
                elif "turno" in k:
                    meta["Turno"] = v
                elif "periodo" in k or "período" in k:
                    meta["Periodo"] = v

    # 4) Buscar la tabla de detalle con columnas de reloj
    candidatos = []
    for df in dfs:
        cols = [str(c).strip().lower() for c in df.columns]
        tiene_fecha = any(c == "fecha" or "fecha" in c for c in cols)
        tiene_entrada = any("entrada" in c for c in cols)
        tiene_salida = any("salida" in c for c in cols)
        if tiene_fecha and (tiene_entrada or tiene_salida):
            candidatos.append(df)

    if not candidatos:
        # Si no hay tabla “obvia”, devolvemos todas crudas
        return exportar_tablas_crudas(dfs)

    # 5) Elegir candidata con más columnas
    detalle_df = max(candidatos, key=lambda d: d.shape[1]).fillna("")

    # 6) Mapeo flexible de nombres de columnas
    def pick(colnames, *opciones):
        low = [str(c).lower().strip() for c in colnames]
        for op in opciones:
            if op in low:
                return colnames[low.index(op)]
        return None

    col_fecha = pick(detalle_df.columns, "fecha")
    col_ent   = pick(detalle_df.columns, "entrada", "hora entrada", "hora de entrada", "ingreso", "hora ingreso")
    col_sal   = pick(detalle_df.columns, "salida", "hora salida", "hora de salida", "egreso", "hora egreso")
    col_desc  = pick(detalle_df.columns, "descripcion", "descripción", "observacion", "observación", "detalle", "comentario")

    # 7) Normalizar filas a nuestro esquema (Fecha dd-mm-aaaa, horas HH:MM:SS)
    registros = []
    for _, row in detalle_df.iterrows():
        fecha_raw = row.get(col_fecha, "")
        ent_raw   = row.get(col_ent, "")
        sal_raw   = row.get(col_sal, "")
        des       = row.get(col_desc, "")

        # Fecha a dd-mm-aaaa
        fecha_str = ""
        try:
            f = pd.to_datetime(str(fecha_raw), dayfirst=True, errors="coerce")
            if pd.notna(f):
                fecha_str = f.strftime("%d-%m-%Y")
        except Exception:
            pass

        registros.append({
            "Fecha": fecha_str or str(fecha_raw),
            "Entrada": _to_hms(ent_raw),
            "Salida": _to_hms(sal_raw),
            "Descripcion": des
        })

    # 8) Armar resultado (igual estilo que Excel real)
    return armar_resultados(registros, meta)

def exportar_tablas_crudas(dfs):
    """
    Devuelve un XLSX con todas las tablas encontradas en el HTML,
    por si el layout cambia y no se puede mapear automáticamente.
    """
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for i, df in enumerate(dfs, 1):
            df.to_excel(writer, sheet_name=f"Tabla_{i}", index=False)
    out.seek(0)
    return out

# =================== UTIL HORA ===================

def _to_hms(x):
    """
    Normaliza formatos de hora comunes a 'HH:MM:SS'.
    Acepta:
      - 'HH:MM:SS'
      - 'HH:MM'
      - 'HMM' / 'HHMM' (ej. 800 -> 08:00:00)
      - otros -> '-'
    """
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none", "-"):
        return "-"
    if re.match(r"^\d{1,2}:\d{2}:\d{2}$", s):
        return s
    if re.match(r"^\d{1,2}:\d{2}$", s):
        return s + ":00"
    if re.match(r"^\d{3,4}$", s):
        s = s.zfill(4)
        return f"{s[:2]}:{s[2:]}:00"
    return "-"
